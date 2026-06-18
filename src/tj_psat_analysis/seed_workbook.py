"""Ingest the seed workbook into canonical roster and enrollment outputs."""

from __future__ import annotations

import csv
import hashlib
import re
import shutil
from collections import Counter
from collections.abc import Iterable
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook

from .constants import (
    CLASS_YEAR_TO_GRADE11_SCHOOL_YEAR,
    PUBLIC_ENROLLMENT_SHEET,
    PUBLIC_ENROLLMENT_SOURCE_TITLE,
    PUBLIC_ENROLLMENT_SOURCE_URL,
    ROSTER_SHEET,
)
from .normalize import aliases_for_school, normalize_school_name, slugify


@dataclass(frozen=True)
class SchoolRecord:
    school_id: str
    pathway: str
    division: str
    sector: str
    school: str
    aliases: tuple[str, ...]
    notes: str
    source_workbook: str
    source_sheet: str
    source_row: int


@dataclass(frozen=True)
class PublicEnrollmentRawRow:
    source_row_id: str
    source_sheet: str
    source_row: int
    school_name_source: str
    state_name: str
    values_by_year: dict[str, object]
    high_school_values_by_year: dict[str, object]


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _string(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _rows_from_sheet(workbook_path: Path, sheet_name: str) -> Iterable[tuple[int, tuple[object, ...]]]:
    workbook = load_workbook(workbook_path, data_only=True, read_only=True)
    worksheet = workbook[sheet_name]
    try:
        yield from enumerate(worksheet.iter_rows(values_only=True), start=1)
    finally:
        workbook.close()


def read_roster(workbook_path: Path, source_workbook_label: str | None = None) -> list[SchoolRecord]:
    rows = list(_rows_from_sheet(workbook_path, ROSTER_SHEET))
    if not rows:
        raise ValueError(f"{ROSTER_SHEET} is empty")

    header = [_string(value) for value in rows[0][1]]
    expected = ["Pathway", "Division", "Sector", "School", "Notes"]
    if header[:5] != expected:
        raise ValueError(f"Unexpected {ROSTER_SHEET} header: {header[:5]}")

    records: list[SchoolRecord] = []
    seen_ids: set[str] = set()
    for row_index, row in rows[1:]:
        values = list(row[:5])
        if not any(_string(value) for value in values):
            continue
        pathway, division, sector, school, notes = [_string(value) for value in values]
        if not school:
            raise ValueError(f"Roster row {row_index} is missing school name")
        base_id = slugify(school)
        school_id = base_id
        suffix = 2
        while school_id in seen_ids:
            school_id = f"{base_id}_{suffix}"
            suffix += 1
        seen_ids.add(school_id)
        records.append(
            SchoolRecord(
                school_id=school_id,
                pathway=pathway,
                division=division,
                sector=sector,
                school=school,
                aliases=tuple(aliases_for_school(school, notes)),
                notes=notes,
                source_workbook=source_workbook_label or str(workbook_path),
                source_sheet=ROSTER_SHEET,
                source_row=row_index,
            )
        )

    tj_rows = [
        record
        for record in records
        if record.school == "Thomas Jefferson High School for Science and Technology"
    ]
    if len(tj_rows) != 1:
        raise ValueError(f"Expected exactly one TJHSST row, found {len(tj_rows)}")
    return records


def _parse_year_from_header(header: str, prefix: str) -> str | None:
    if not header.startswith(prefix):
        return None
    match = re.search(r"(\d{4}-\d{2})$", header)
    return match.group(1) if match else None


def read_public_enrollment_raw(workbook_path: Path) -> list[PublicEnrollmentRawRow]:
    rows = list(_rows_from_sheet(workbook_path, PUBLIC_ENROLLMENT_SHEET))
    header_position = None
    header: list[str] = []
    for position, (_, row) in enumerate(rows):
        header_values = [_string(value) for value in row]
        if header_values and header_values[0] == "School Name":
            header_position = position
            header = header_values
            break
    if header_position is None:
        raise ValueError(f"Could not find public enrollment header in {PUBLIC_ENROLLMENT_SHEET}")

    grade11_columns: dict[int, str] = {}
    grades9_12_columns: dict[int, str] = {}
    for index, label in enumerate(header):
        year = _parse_year_from_header(label, "Grade 11 Students [Public School]")
        if year:
            grade11_columns[index] = year
        year = _parse_year_from_header(label, "Grades 9-12 Students [Public School]")
        if year:
            grades9_12_columns[index] = year

    records: list[PublicEnrollmentRawRow] = []
    for row_index, row in rows[header_position + 1 :]:
        row_values: list[object] = list(row)
        school_name = _string(row_values[0] if row_values else "")
        if not school_name:
            continue
        if school_name.startswith("Data Source:"):
            break
        state_name = _string(row_values[1] if len(row_values) > 1 else "")
        grade11: dict[str, object] = {
            year: row_values[index] if index < len(row_values) else None
            for index, year in grade11_columns.items()
        }
        grades9_12: dict[str, object] = {
            year: row_values[index] if index < len(row_values) else None
            for index, year in grades9_12_columns.items()
        }
        records.append(
            PublicEnrollmentRawRow(
                source_row_id=f"sheet6_row_{row_index:03d}",
                source_sheet=PUBLIC_ENROLLMENT_SHEET,
                source_row=row_index,
                school_name_source=school_name,
                state_name=state_name,
                values_by_year=grade11,
                high_school_values_by_year=grades9_12,
            )
        )
    return records


def parse_nces_value(value: object) -> tuple[int | None, str]:
    if value is None or value == "":
        return None, "blank"
    if isinstance(value, (int, float)):
        return int(value), "reported"
    text = str(value).strip()
    if text == "\u2020":
        return None, "not_applicable"
    if text in {"-", "\u2013"}:
        return None, "missing"
    if text == "\u2021":
        return None, "failed_quality_standard"
    try:
        return int(float(text.replace(",", ""))), "reported"
    except ValueError:
        return None, f"unparsed:{text}"


def _public_enrollment_index(
    raw_rows: list[PublicEnrollmentRawRow],
) -> dict[str, list[PublicEnrollmentRawRow]]:
    index: dict[str, list[PublicEnrollmentRawRow]] = {}
    for row in raw_rows:
        key = normalize_school_name(row.school_name_source)
        index.setdefault(key, []).append(row)
    return index


def _candidate_keys(record: SchoolRecord) -> list[str]:
    keys: list[str] = []
    for alias in record.aliases:
        key = normalize_school_name(alias)
        if key and key not in keys:
            keys.append(key)
    return keys


def match_public_enrollment_rows(
    roster: list[SchoolRecord],
    raw_rows: list[PublicEnrollmentRawRow],
) -> dict[str, tuple[str, list[PublicEnrollmentRawRow]]]:
    index = _public_enrollment_index(raw_rows)
    matches: dict[str, tuple[str, list[PublicEnrollmentRawRow]]] = {}
    for record in roster:
        if record.sector != "Public":
            matches[record.school_id] = ("not_public_school", [])
            continue
        candidates: list[PublicEnrollmentRawRow] = []
        for key in _candidate_keys(record):
            for row in index.get(key, []):
                if row not in candidates:
                    candidates.append(row)
        if not candidates:
            matches[record.school_id] = ("source_row_not_found", [])
        elif len(candidates) == 1:
            matches[record.school_id] = ("matched", candidates)
        else:
            matches[record.school_id] = ("ambiguous_source_name", candidates)
    return matches


def roster_to_rows(roster: list[SchoolRecord]) -> list[dict[str, object]]:
    return [
        {
            "school_id": record.school_id,
            "pathway": record.pathway,
            "division": record.division,
            "sector": record.sector,
            "school": record.school,
            "aliases": "|".join(record.aliases),
            "notes": record.notes,
            "source_workbook": record.source_workbook,
            "source_sheet": record.source_sheet,
            "source_row": record.source_row,
        }
        for record in roster
    ]


def public_raw_to_rows(raw_rows: list[PublicEnrollmentRawRow]) -> list[dict[str, object]]:
    rows: list[dict[str, object]] = []
    for raw in raw_rows:
        for school_year, raw_value in raw.values_by_year.items():
            value, status = parse_nces_value(raw_value)
            grades9_12, grades9_12_status = parse_nces_value(raw.high_school_values_by_year.get(school_year))
            rows.append(
                {
                    "source_row_id": raw.source_row_id,
                    "source_sheet": raw.source_sheet,
                    "source_row": raw.source_row,
                    "school_name_source": raw.school_name_source,
                    "state_name": raw.state_name,
                    "school_year": school_year,
                    "grade11_students": value,
                    "grade11_status": status,
                    "grades9_12_students": grades9_12,
                    "grades9_12_status": grades9_12_status,
                }
            )
    return rows


def public_enrollment_long_rows(
    roster: list[SchoolRecord],
    raw_rows: list[PublicEnrollmentRawRow],
    workbook_hash: str,
) -> list[dict[str, object]]:
    matches = match_public_enrollment_rows(roster, raw_rows)
    rows: list[dict[str, object]] = []
    for record in roster:
        if record.sector != "Public":
            continue
        match_status, candidates = matches[record.school_id]
        for class_year, school_year in CLASS_YEAR_TO_GRADE11_SCHOOL_YEAR.items():
            output = {
                "school_id": record.school_id,
                "school": record.school,
                "class_year": class_year,
                "grade11_school_year": school_year,
                "grade11_enrollment": "",
                "enrollment_status": match_status,
                "enrollment_source_title": PUBLIC_ENROLLMENT_SOURCE_TITLE,
                "enrollment_source_url": PUBLIC_ENROLLMENT_SOURCE_URL,
                "enrollment_source_date": "",
                "enrollment_source_hash": workbook_hash,
                "enrollment_source_sheet": PUBLIC_ENROLLMENT_SHEET,
                "enrollment_source_rows": ";".join(candidate.source_row_id for candidate in candidates),
            }
            if school_year == "2024-25":
                output["enrollment_status"] = "source_year_not_in_seed"
                output["enrollment_source_rows"] = ""
            elif match_status == "matched":
                raw = candidates[0]
                value, status = parse_nces_value(raw.values_by_year.get(school_year))
                output["grade11_enrollment"] = value if value is not None else ""
                output["enrollment_status"] = status
                output["enrollment_source_rows"] = raw.source_row_id
            rows.append(output)
    return rows


def class_year_mapping_rows() -> list[dict[str, object]]:
    return [
        {
            "class_year": class_year,
            "qualifying_psat_year": f"Fall {class_year - 2}",
            "grade11_school_year": grade11_school_year,
        }
        for class_year, grade11_school_year in CLASS_YEAR_TO_GRADE11_SCHOOL_YEAR.items()
    ]


def seed_panel_rows(
    roster: list[SchoolRecord],
    public_enrollment_rows: list[dict[str, object]],
) -> list[dict[str, object]]:
    public_lookup = {
        (str(row["school_id"]), int(str(row["class_year"]))): row for row in public_enrollment_rows
    }
    panel: list[dict[str, object]] = []
    for record in roster:
        for class_year, school_year in CLASS_YEAR_TO_GRADE11_SCHOOL_YEAR.items():
            public_row = public_lookup.get((record.school_id, class_year))
            if public_row:
                grade11_enrollment = public_row["grade11_enrollment"]
                enrollment_status = public_row["enrollment_status"]
                enrollment_source_title = public_row["enrollment_source_title"]
                enrollment_source_url = public_row["enrollment_source_url"]
                enrollment_source_date = public_row["enrollment_source_date"]
                enrollment_source_hash = public_row["enrollment_source_hash"]
                enrollment_source_rows = public_row["enrollment_source_rows"]
            elif record.sector == "Private":
                grade11_enrollment = ""
                enrollment_status = "private_pss_not_ingested"
                enrollment_source_title = ""
                enrollment_source_url = ""
                enrollment_source_date = ""
                enrollment_source_hash = ""
                enrollment_source_rows = ""
            else:
                grade11_enrollment = ""
                enrollment_status = "public_enrollment_not_ingested"
                enrollment_source_title = ""
                enrollment_source_url = ""
                enrollment_source_date = ""
                enrollment_source_hash = ""
                enrollment_source_rows = ""
            panel.append(
                {
                    "school_id": record.school_id,
                    "school": record.school,
                    "pathway": record.pathway,
                    "division": record.division,
                    "sector": record.sector,
                    "class_year": class_year,
                    "grade11_school_year": school_year,
                    "nmsf_count": "",
                    "nmsf_status": "source_pending",
                    "nmsf_source_title": "",
                    "nmsf_source_url": "",
                    "nmsf_source_date": "",
                    "nmsf_source_hash": "",
                    "nmsf_source_scope": "",
                    "grade11_enrollment": grade11_enrollment,
                    "enrollment_status": enrollment_status,
                    "enrollment_source_title": enrollment_source_title,
                    "enrollment_source_url": enrollment_source_url,
                    "enrollment_source_date": enrollment_source_date,
                    "enrollment_source_hash": enrollment_source_hash,
                    "enrollment_source_rows": enrollment_source_rows,
                    "nmsf_per_100_grade11": "",
                    "notes": record.notes,
                }
            )
    return panel


def _markdown_table(headers: list[str], rows: list[list[object]]) -> str:
    rendered = [
        "| " + " | ".join(headers) + " |",
        "| " + " | ".join("---" for _ in headers) + " |",
    ]
    for row in rows:
        rendered.append("| " + " | ".join(str(value) for value in row) + " |")
    return "\n".join(rendered)


def _counter_rows(counter: Counter[str]) -> list[list[object]]:
    return [[key or "(blank)", value] for key, value in sorted(counter.items())]


def _format_review_rows(rows: list[list[object]], empty_message: str) -> str:
    if not rows:
        return empty_message
    return _markdown_table(["School", "Review status", "Source rows", "Source names"], rows)


def _roster_missing_required_fields(roster: list[SchoolRecord]) -> list[list[object]]:
    required = ("school_id", "pathway", "division", "sector", "school")
    missing_rows: list[list[object]] = []
    for record in roster:
        missing = [field for field in required if not str(getattr(record, field, "")).strip()]
        if missing:
            missing_rows.append([record.source_row, record.school or "(blank)", ", ".join(missing)])
    return missing_rows


def _duplicate_rows_by_normalized_name(roster: list[SchoolRecord]) -> list[list[object]]:
    grouped: dict[str, list[SchoolRecord]] = {}
    for record in roster:
        key = normalize_school_name(record.school)
        grouped.setdefault(key, []).append(record)

    duplicates: list[list[object]] = []
    for key, records in sorted(grouped.items()):
        if len(records) > 1:
            duplicates.append(
                [
                    key,
                    len(records),
                    "; ".join(record.school for record in records),
                    "; ".join(record.school_id for record in records),
                ]
            )
    return duplicates


def _raw_enrollment_duplicate_rows(raw_rows: list[PublicEnrollmentRawRow]) -> list[list[object]]:
    grouped: dict[str, list[PublicEnrollmentRawRow]] = {}
    for row in raw_rows:
        grouped.setdefault(normalize_school_name(row.school_name_source), []).append(row)

    duplicates: list[list[object]] = []
    for key, rows in sorted(grouped.items()):
        if len(rows) > 1:
            duplicates.append(
                [
                    key,
                    len(rows),
                    "; ".join(row.source_row_id for row in rows),
                    "; ".join(row.school_name_source for row in rows),
                ]
            )
    return duplicates


def _enrollment_review_rows(
    roster: list[SchoolRecord],
    raw_rows: list[PublicEnrollmentRawRow],
) -> list[list[object]]:
    matches = match_public_enrollment_rows(roster, raw_rows)
    review_rows: list[list[object]] = []
    for record in roster:
        if record.sector != "Public":
            continue
        status, candidates = matches[record.school_id]
        if status in {"ambiguous_source_name", "source_row_not_found"}:
            review_rows.append(
                [
                    record.school,
                    status,
                    "; ".join(row.source_row_id for row in candidates) or "(none)",
                    "; ".join(row.school_name_source for row in candidates) or "(none)",
                ]
            )
    return review_rows


def _workbook_sheet_names(workbook_path: Path) -> list[str]:
    workbook = load_workbook(workbook_path, read_only=True)
    sheet_names = list(workbook.sheetnames)
    workbook.close()
    return sheet_names


def _display_path(path: Path) -> str:
    try:
        return str(path.relative_to(Path.cwd()))
    except ValueError:
        return str(path)


def build_workbook_ingestion_report(
    workbook_path: Path,
    workbook_hash: str,
    roster: list[SchoolRecord],
    public_raw: list[PublicEnrollmentRawRow],
    public_long: list[dict[str, object]],
    panel: list[dict[str, object]],
    manual_workbook_copy: Path | None = None,
) -> str:
    sheet_names = _workbook_sheet_names(workbook_path)
    used_sheets = [ROSTER_SHEET, PUBLIC_ENROLLMENT_SHEET]
    ignored_sheets = [sheet for sheet in sheet_names if sheet not in used_sheets]
    numeric_nmsf_rows = [row for row in panel if str(row.get("nmsf_count", "")).strip()]

    lines = [
        "# Workbook Ingestion Data Quality Report",
        "",
        "This report is generated from the seed workbook parser. It is deterministic",
        "and records review queues rather than guessing through ambiguous data.",
        "",
        "## Source workbook",
        "",
        _markdown_table(
            ["Field", "Value"],
            [
                ["Workbook path", _display_path(workbook_path)],
                ["Workbook SHA-256", workbook_hash],
                [
                    "Manual copy path",
                    _display_path(manual_workbook_copy) if manual_workbook_copy else "(not requested)",
                ],
                ["Sheets read", ", ".join(used_sheets)],
                ["Sheets ignored", ", ".join(ignored_sheets) or "(none)"],
            ],
        ),
        "",
        "The parser reads only `raw` and `Sheet6`. It does not read values from",
        "`nsmf 2019`, and the generated seed panel contains no numeric NMSF counts.",
        "",
        "## Output summary",
        "",
        _markdown_table(
            ["Output", "Rows"],
            [
                ["canonical school roster", len(roster)],
                ["public enrollment raw long rows", len(public_raw_to_rows(public_raw))],
                ["public grade-11 enrollment rows", len(public_long)],
                ["class-year mapping rows", len(CLASS_YEAR_TO_GRADE11_SCHOOL_YEAR)],
                ["seed panel rows", len(panel)],
                ["numeric NMSF rows in seed panel", len(numeric_nmsf_rows)],
            ],
        ),
        "",
        "## Roster coverage",
        "",
        _markdown_table(["Sector", "Schools"], _counter_rows(Counter(row.sector for row in roster))),
        "",
        _markdown_table(["Pathway", "Schools"], _counter_rows(Counter(row.pathway for row in roster))),
        "",
        "## Enrollment status counts",
        "",
        _markdown_table(
            ["Enrollment status", "Rows"],
            _counter_rows(Counter(str(row["enrollment_status"]) for row in public_long)),
        ),
        "",
        "## Missing required roster fields",
        "",
    ]

    missing_required = _roster_missing_required_fields(roster)
    if missing_required:
        lines.append(_markdown_table(["Workbook row", "School", "Missing fields"], missing_required))
    else:
        lines.append("No missing required roster fields.")

    duplicate_roster_rows = _duplicate_rows_by_normalized_name(roster)
    lines.extend(["", "## Duplicate roster names", ""])
    if duplicate_roster_rows:
        lines.append(
            _markdown_table(
                ["Normalized name", "Rows", "Schools", "School IDs"],
                duplicate_roster_rows,
            )
        )
    else:
        lines.append("No duplicate normalized roster school names.")

    duplicate_source_rows = _raw_enrollment_duplicate_rows(public_raw)
    lines.extend(["", "## Duplicate public enrollment source names", ""])
    if duplicate_source_rows:
        lines.append(
            _markdown_table(
                ["Normalized source name", "Rows", "Source row IDs", "Source names"],
                duplicate_source_rows,
            )
        )
    else:
        lines.append("No duplicate normalized public enrollment source names.")

    lines.extend(["", "## Enrollment review queue", ""])
    lines.append(
        _format_review_rows(
            _enrollment_review_rows(roster, public_raw),
            "No ambiguous or missing public enrollment source-name matches.",
        )
    )

    lines.extend(
        [
            "",
            "## NMSF exclusion check",
            "",
            _markdown_table(
                ["Check", "Result"],
                [
                    ["`nsmf 2019` read by parser", "no"],
                    ["Numeric NMSF rows in seed panel", len(numeric_nmsf_rows)],
                    ["Default NMSF status", "source_pending"],
                ],
            ),
            "",
        ]
    )
    return "\n".join(lines)


def write_csv(path: Path, rows: list[dict[str, object]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    fieldnames = list(rows[0].keys()) if rows else []
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames, lineterminator="\n")
        writer.writeheader()
        writer.writerows(rows)


def write_text(path: Path, text: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(text, encoding="utf-8")


def ensure_manual_workbook_copy(workbook_path: Path, manual_dir: Path) -> Path:
    manual_dir.mkdir(parents=True, exist_ok=True)
    target = manual_dir / workbook_path.name
    if target.exists():
        if sha256_file(target) != sha256_file(workbook_path):
            raise ValueError(f"{target} exists but is not byte-identical to {workbook_path}")
        return target
    shutil.copy2(workbook_path, target)
    return target


def build_seed_outputs(
    workbook_path: Path,
    output_dir: Path,
    processed_dir: Path | None = None,
    report_dir: Path | None = None,
    manual_dir: Path | None = None,
) -> dict[str, Path]:
    workbook_path = workbook_path.resolve()
    output_dir.mkdir(parents=True, exist_ok=True)

    workbook_hash = sha256_file(workbook_path)
    manual_workbook_copy = ensure_manual_workbook_copy(workbook_path, manual_dir) if manual_dir else None
    try:
        source_workbook_label = str(workbook_path.relative_to(Path.cwd()))
    except ValueError:
        source_workbook_label = str(workbook_path)
    roster = read_roster(workbook_path, source_workbook_label)
    public_raw = read_public_enrollment_raw(workbook_path)
    public_long = public_enrollment_long_rows(roster, public_raw, workbook_hash)
    panel = seed_panel_rows(roster, public_long)

    outputs = {
        "canonical_schools": output_dir / "canonical_schools.csv",
        "public_enrollment_raw": output_dir / "public_enrollment_raw.csv",
        "public_grade11_enrollment": output_dir / "public_grade11_enrollment.csv",
        "panel_seed": output_dir / "panel_seed.csv",
    }
    write_csv(outputs["canonical_schools"], roster_to_rows(roster))
    write_csv(outputs["public_enrollment_raw"], public_raw_to_rows(public_raw))
    write_csv(outputs["public_grade11_enrollment"], public_long)
    write_csv(outputs["panel_seed"], panel)

    if processed_dir:
        processed_outputs = {
            "schools": processed_dir / "schools.csv",
            "public_enrollment": processed_dir / "public_enrollment.csv",
            "class_year_mapping": processed_dir / "class_year_mapping.csv",
        }
        write_csv(processed_outputs["schools"], roster_to_rows(roster))
        write_csv(processed_outputs["public_enrollment"], public_long)
        write_csv(processed_outputs["class_year_mapping"], class_year_mapping_rows())
        outputs.update(processed_outputs)

    if report_dir:
        report_path = report_dir / "workbook_ingestion.md"
        write_text(
            report_path,
            build_workbook_ingestion_report(
                workbook_path=workbook_path,
                workbook_hash=workbook_hash,
                roster=roster,
                public_raw=public_raw,
                public_long=public_long,
                panel=panel,
                manual_workbook_copy=manual_workbook_copy,
            ),
        )
        outputs["workbook_ingestion_report"] = report_path

    if manual_workbook_copy:
        outputs["manual_workbook"] = manual_workbook_copy

    return outputs
